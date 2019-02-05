import asyncio
import discord
import random
import traceback
from discord.ext import commands
from discord.ext.commands import Bot
import time
from random import choice
import youtube_dl
from urllib.request import urlopen, Request
import urllib
from urllib.request import Request
import urllib.request
import os
from discord import Member
import sys
import json
import args
import player
import bs4
import datetime
import time
import requests
import ctx
from selenium import webdriver
import openpyxl

client = discord.Client()


countG = 0
players = {}


# 봇이 구동되었을 때 동작되는 코드입니다.
@client.event
async def on_ready():
    print("Logged in as ")  # 화면에 봇의 아이디, 닉네임이 출력됩니다.
    print(client.user.name)
    print(client.user.id)
    print("===========")
    # 디스코드에는 현재 본인이 어떤 게임을 플레이하는지 보여주는 기능이 있습니다.
    # 이 기능을 사용하여 봇의 상태를 간단하게 출력해줄 수 있습니다.
    await client.change_presence(game=discord.Game(name="~도움말", type=1))

# 봇이 새로운 메시지를 수신했을때 동작되는 코드입니다.
@client.event
async def on_message(message):
    if message.author.bot:  # 만약 메시지를 보낸사람이 봇일 경우에는
        return None  # 동작하지 않고 무시합니다.

    id = message.author.id  # id라는 변수에는 메시지를 보낸사람의 ID를 담습니다.
    channel = message.channel  # channel이라는 변수에는 메시지를 받은 채널의 ID를 담습니다.


    if message.content.startswith("~도움말"):
        await client.send_message(channel, '```셔노봇은 연기 스터디, 대본 리딩에 적극적으로 활용 가능한 디스코드 봇입니다```')


    if message.content.startswith("~도움말"):
        channel = message.channel
        embed = discord.Embed(
            title='셔노봇 명령어 리스트',
            description=':arrow_down:',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='END')
        embed.add_field(name='스터디 카테고리', value='**~단문**\n``랜덤으로 단문 하나를 지정해 채팅창에 띄웁니다.``\n'
                                               '**~대본모음**\n``대본을 업로드해주는 사이트들의 링크 목록을 작성해 채팅창에 띄웁니다``\n'
                                               '**~인원별대본**\n``명령어 입력, 띄어쓰기 후 원하는 인원을 입력하면 해당 인원수에 맞춰 대본 링크 목록들을 올려줍니다. (최대 9명까지)``\n'
                                               '**~공채대본 남/여**\n``가장 최근 공지됐던 방송사의 공채 대본을 올려줍니다. (해당 기능은 새로운 공채시험이 공지될때마다 업데이트 됩니다.)``\n', inline=False)
        embed.add_field(name='게임 카테고리', value='\n**~주사위**\n``주사위를 던지며 결과를 알려줍니다.``\n'
                                              '**~골라**\n``명령어 입력, 띄어쓰기 후 원하는 항목들을 입력하면 입력한 항목들 중에 하나를 랜덤으로 골라줍니다.``\n```ex. ~골라 치킨 햄버거```'
                                              '**~사다리타기**\n``명령어 입력, 띄어쓰기 후 게임에 참여할 항목들, 결과 항목들을 입력하면 사다리타기를 시작합니다.``\n```ex. ~사다리타기 철수 영희 지민/1팀 2팀 3팀```'
                                              '**~제비뽑기**\n``명령어 입력, 띄어쓰기 후 원하는 제비의 갯수를 입력하면 지정 갯수만큼 제비뽑기 결과를 알려줍니다.``\n', inline=False)
        embed.add_field(name='커맨드 카테고리', value='\n**~커맨드**\n``명령어 입력, 띄어쓰기 후 원하는 단어와 원하는 대답을 입력한 후 추가한 단어를 그대로 입력하면 봇이 추가된 대답을 해줍니다.``\n```ex. ~커맨드 안녕/반가워!```'
                                               '**~커맨드목록**\n``해당 명령어를 입력하면 자신이 추가했던 커맨드들을 목록으로 나열하여 올려줍니다.``\n'
                                               '**~커맨드삭제**\n``명령어 입력, 띄어쓰기 후 추가했던 단어를 입력하면 해당 커맨드가 삭제됩니다.``\n```ex. ~커맨드삭제 안녕```', inline=False)
        embed.add_field(name='기본기능 카테고리', value='\n**~이미지**\n``명령어 입력, 띄어쓰기 후 원하는 검색어를 입력하면 해당 검색어의 이미지를 랜덤으로 뽑아 링크와 함께 올려줍니다.``\n```검색엔진 - NAVER```'
                                                '**~비디오**\n``명령어 입력, 띄어쓰기 후 원하는 검색어를 입력하면 해당 검색어의 영상을 가져와 링크와 함께 올려줍니다.``\n```검색엔진 - YOUTUBE```'
                                                '**~날씨**\n``명령어 입력, 띄어쓰기 후 날씨를 알고 싶은 지역을 입력하면 해당 지역의 날씨와 대기 상태를 알려줍니다.``\n```ex. ~날씨 서울```'
                                                '**~청소**\n``명령어 입력 후 지우고 싶은 메세지의 개수를 입력하면 해당 개수만큼의 메세지를 삭제시킵니다.``\n'
                                                '**~봇초대**\n``해당 명령어를 입력하면 셔노봇을 초대할 수 있는 링크를 띄워줍니다.``\n', inline=False)
        embed.add_field(name='음악기능 카테고리', value='\n**~들어와**\n``셔노봇이 음성채널에 들어옵니다.``\n'
                                                '**~나가**\n``셔노봇이 음성채널에서 나갑니다.``\n'
                                                '**~재생**\n``명령어 입력, 띄어쓰기 후 재생시키고 싶은 영상의 제목을 입력하면 해당 영상의 음원을 재생해줍니다.``\n```반드시 "~들어와" 명령어 입력 후 사용할 것```'
                                                '**~일시정지**\n``현재 틀어져있는 곡의 재생을 일시정지시킵니다.``\n'
                                                '**~다시재생**\n``일시정지 해놓은 곡을 다시 재생시킵니다.``\n'
                                                '**~재생목록**\n``예약 후 대기중인 곡의 목록을 확인합니다.``\n```예약기능은 "~재생" 명령어를 통해서 사용```'
                                                '**~재생확인**\n``현재 재생중인 곡의 정보를 확인합니다.``\n'
                                                '**~스킵**\n``현재 재생중인 곡을 스킵합니다.``\n'
                                                '**~예약취소**\n``현재 예약목록에 추가된 모든 곡을 초기화시킵니다.``\n'
                                                '**~음량**\n``현재 설정돼있는 봇의 기본 음량을 확인합니다.``\n', inline=False)

        await client.send_message(channel, embed=embed)

        if message.content.startswith("~도움말"):
            await client.send_message(channel, '```셔노봇 공식 디스코드 서버 링크 - https://discord.gg/7btftpd```')



    if message.content.startswith("~대본모음"):
        channel = message.channel
        embed = discord.Embed(
            title='대본 링크 리스트',
            description=':arrow_down:',
            colour=discord.Colour.green()
        )

        embed.set_footer(text='END')
        embed.add_field(name='<변기철> : KBS라디오드라마 / 애니메이션 / 드라마 / 영화', value='http://blog.naver.com/smtm3287',
                        inline=False)
        embed.add_field(name='<두더지> : KBS라디오드라마 / 음악에세이 / 애니메이션 / 나레이션 단문',
                        value='http://kimjm781.blog.me/220279894757', inline=False)
        embed.add_field(name='<싸이월드 클럽 "성우가 되자"> : KBS라디오드라마', value='http://club.cyworld.com/ClubV1/Home.cy/50418311',
                        inline=False)
        embed.add_field(name='<이혜> : KBS라디오드라마', value='https://blog.naver.com/sbs_____', inline=False)
        embed.add_field(name='<홍> : 음악에세이', value='http://blog.naver.com/hsh0364', inline=False)
        embed.add_field(name='<위대한 인성님> : 애니메이션 / 영화 / 드라마', value='https://blog.naver.com/migif1124', inline=False)
        embed.add_field(name='<루카스> : 애니메이션 / 영화 / 드라마', value='https://blog.naver.com/lukasix', inline=False)
        embed.add_field(name='<호빵이> : 애니메이션', value='http://blog.daum.net/kehstudent/13620272', inline=False)
        embed.add_field(name='<호피> : 애니메이션 / 영화 / 드라마', value='http://stephanos_.blog.me/30145688436', inline=False)
        embed.add_field(name='<향연> : 애니메이션 / 영화 / 드라마', value='http://only_dear.blog.me/140179808557', inline=False)
        embed.add_field(name='<vat19 우> : 애니메이션 / 영화 / 드라마', value='http://blog.naver.com/vat19', inline=False)
        embed.add_field(name='<네다네> : 애니메이션 / 영화 / 드라마 / KBS무대', value='https://blog.naver.com/tjsdlr123', inline=False)
        embed.add_field(name='<너부리> : 애니메이션 / 영화 / 동화', value='http://blog.daum.net/dhhsdhdfhdrthr', inline=False)
        embed.add_field(name='<옴므> : 애니메이션 / 영화', value='https://blog.naver.com/bbyong456', inline=False)
        embed.add_field(name='<뀨뀨라에일> : 애니메이션', value='https://blog.naver.com/ceojhshin', inline=False)
        embed.add_field(name='<글쟁이 엄석대> : 애니메이션 / KBS라디오 드라마 / 영화', value='https://blog.naver.com/kwoo0221',
                        inline=False)
        embed.add_field(name='<빡구> : 대본 링크 모음집', value='https://blog.naver.com/dyekwjddus12/220465259330', inline=False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith("~인원별대본 2"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <2명>',
            description=':arrow_down:',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[1남 1공]', value='http://blog.naver.com/dudgus819/220959840794 [어느 연금 술사의 편지]', inline=False)
        embed.add_field(name='[2남]', value='http://blog.naver.com/ghfkddl67/220319134100 우리 집의 욕조사정 1화\n'
                                           'http://blueattack.blog.me/220005500002 셜록과 마이크로프트\n'
                                           'http://blueattack.blog.me/30155206559 셜록과 짐 모리아티\n'
                                           'http://blueattack.blog.me/30069493853 더파이팅 챔피언 로드\n'
                                           'http://blueattack.blog.me/30086174304 더파이팅 사형집행\n'
                                           'http://blog.naver.com/vat19/220973311041 셜록 E202 http://blog.naver.com/snowrain1100/221069124285 이누야샤 극장판 3기 천하패도의 검\n'
                                           'http://blog.naver.com/dudgus819/220936450547  [어노잉 오렌지 사과편] <ㅡ 약\n'
                                           'http://blog.naver.com/dudgus819/220937706040  [조커 Why so serious?]\n'
                                           'http://blog.naver.com/dudgus819/220990328841  [디아블로 3 말티엘 영혼을 거두는 자]', inline=False)
        embed.add_field(name='[1남 1녀]', value='http://blog.naver.com/snowrain1100/221061096546 오빠를 고칠 약은 없다 1화\n'
                                              'http://blog.naver.com/snowrain1100/221077595141 소드아트 온라인 2기12화\n'
                                              'http://blog.naver.com/snowrain1100/221157105801 살육의 천사 단편만화\n'
                                              'https://blog.naver.com/snowrain1100/221314345774 살육의 천사 1화\n'
                                              'http://blog.naver.com/bbyong456/220549460825\n'
                                              'http://blog.naver.com/bbyong456/220549564909\n'
                                              'http://blog.naver.com/bbyong456/220552346127\n'
                                              '월간소녀 노자키군 Ep.1 http://blog.naver.com/bbyong456/220309693462\n'
                                              '아라카와 언더 더 브릿지 Ep.1 http://blog.naver.com/bbyong456/220303205112\n'
                                              '더파이팅 챔피언 로드(민시경 인터뷰) http://blueattack.blog.me/30086123914\n'
                                              '페이트 제로 http://blueattack.blog.me/30138039425\n'
                                              '뷰티 인사이드 http://blog.naver.com/teajin288/220558027385\n'
                                              '해를 품은 달(1-1) http://blog.naver.com/tmt234/110135572377\n'
                                              '월간순정 노자키군(2-2) http://blog.naver.com/dpwl970419/220210995487\n'
                                              '월간순정 노자키군 (1-1) http://blog.naver.com/dpwl970419/220209767179\n'
                                              '월간순정 노자키군(1-2) http://blog.naver.com/dpwl970419/220209800076', inline=False)
        embed.add_field(name='[2녀]', value='마리아홀릭 0101 http://blog.naver.com/allstar0_7/50078674996', inline=False)
        embed.add_field(name='[1녀 1공]', value='A.I-4 http://stephanos_.blog.me/30145834272\n엑스 0903 http://blog.naver.com/allstar0_7/50078310430', inline=False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith("~인원별대본 3"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <3명>',
            description=':arrow_down:',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[2남 1녀]', value='http://blog.naver.com/dudgus819/220936076951  [스타크래프트 공허의유산 2]\n'
                                              'http://blog.naver.com/dudgus819/221412942343   [오버워치 - 재회(맥크리)]\n'
                                              'http://blog.naver.com/vat19/220814826851 - 혼밥하는 만화 [1남/1녀/1공]\n'
                                              'http://blog.naver.com/vat19/220798094794 - 오버워치 시네마틱 용 [3남]\n'
                                              'http://blog.naver.com/vat19/220820150122 - 주토피아 3 [2남/1녀]\n'
                                              'ttp://blog.naver.com/vat19/220932110555 - 캐릭캐릭 체인지 예고편 [1남/1녀/1공]\n'
                                              'http://blog.naver.com/vat19/220932183298 - 사카모토입니다만? [3남]\n'
                                              'http://blog.naver.com/vat19/220950972200 - 화이 2 [3남]\n'
                                              'http://blog.naver.com/vat19/220976889396 - 썸남(웹드) 2화 [2남/1녀]', inline=False)
        embed.add_field(name='[3남]', value='돌격!크로마티고교 Ep.6 http://blog.naver.com/bbyong456/220548812208\n'
                                           '돌격!크로마티고교 Ep.4 http://blog.naver.com/bbyong456/220061098885\n'
                                           '원펀맨 Ep.1-2 http://blog.naver.com/bbyong456/220540055716\n'
                                           '엑스 0201 http://blog.naver.com/allstar0_7/50078234886\n'
                                           'http://blog.naver.com/dudgus819/221288851725 - [장삐쭈 단편선 –욕]', inline=False)
        embed.add_field(name='[2남 1녀]', value='원펀맨 Ep.2-1 http://blog.naver.com/bbyong456/220544111983\n'
                                              '요괴워치 1화-2 http://blog.naver.com/teajin288/220581073594\n'
                                              '아라카와 언더 더 브릿지 Ep.2 http://blog.naver.com/bbyong456/220314171977\n'
                                              '나츠메우인장 0102 http://blog.naver.com/allstar0_7/50079164595\n'
                                              '월간소녀 노자키군 Ep.2 http://blog.naver.com/bbyong456/220317461702\n'
                                              '바다가 들린다 http://blog.naver.com/ghfkddl67/220531899338\n'
                                              '노블레스 http://blog.naver.com/cksdnda241/220668693151\n'
                                              '엑스 0303 http://blog.naver.com/allstar0_7/50078270366\n'
                                              '원피스 2화 http://blueattack.blog.me/30156491411\n'
                                              '주군의 태양 3회-1 http://stephanos_.blog.me/30177313066\n'
                                              '주군의 태양 2회-1 http://stephanos_.blog.me/30177094922\n'
                                              '도쿄구울(1-1) http://blog.naver.com/dpwl970419/220209955890\n'
                                              '월간순정 노자키군(6-2) http://blog.naver.com/dpwl970419/220349948559\n'
                                              '월간순정 노자키군(7) http://blog.naver.com/dpwl970419/220350166076\n'
                                              '월간순정 노자키군(3) http://blog.naver.com/dpwl970419/220228540951\n'
                                              '월간순정 노자키군(8-1) http://blog.naver.com/dpwl970419/220350410064\n'
                                              'http://blog.naver.com/snowrain1100/221052681658 소드아트온라인 1기25화\n'
                                              'http://blog.naver.com/snowrain1100/221090938983 클로저스 "티나" 오디오 무비', inline=False)
        embed.add_field(name='[2남 1공]', value='마기(1-1) http://blog.naver.com/dpwl970419/220216651487', inline=False)
        embed.add_field(name='[1남 2녀]', value='경계의 저편 http://blog.naver.com/ghfkddl67/220324014257\n'
                                              '월간소녀 노자키군 Ep.2.5 http://blog.naver.com/bbyong456/220320350100\n'
                                              '보노보노 Ep.1 http://blog.naver.com/bbyong456/220056050624\n'
                                              '보노보노 Ep.2 http://blog.naver.com/bbyong456/220308591240\n'
                                              '이웃집 토토로 1부 http://blog.naver.com/ghfkddl67/220502125797\n'
                                              '너에게 닿기를 1화 http://blueattack.blog.me/30137543207\n'
                                              '매지컬 고삼즈(1~3) http://blog.naver.com/dpwl970419/220368014116\n'
                                              '매지컬 고삼즈(4~6) http://blog.naver.com/dpwl970419/220369048848\n'
                                              '월간순정 노자키군 (2-2) http://blog.naver.com/dpwl970419/220211107689', inline=False)
        embed.add_field(name='[1남 1녀 1공]', value='마기(2-1) http://blog.naver.com/dpwl970419/220216653413', inline=False)
        embed.add_field(name='[1남 2공]', value='헌터X헌터 Ep.4 http://blog.naver.com/bbyong456/220614175213', inline=False)
        embed.add_field(name='[1녀 2공]', value='뽀로로 http://blog.naver.com/ghfkddl67/220328804243', inline=False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith("~인원별대본 4"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <4명>',
            description=':arrow_down:',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[2남 2녀]', value='http://blog.naver.com/dudgus819/220932002442   [오버워치 -영웅 (솔져76)]\n'
                                              'http://blog.naver.com/dudgus819/221412942343   [오버워치 - 재회(맥크리,애쉬)]',
                        inline=False)
        embed.add_field(name='[2남 1녀1공]', value='http://blog.naver.com/dudgus819/220938096561 [디아블로 3 ]\n'
                                                'http://blog.naver.com/dudgus819/221064624953 [보스베이비 -4]', inline=False)
        embed.add_field(name='[3남 1녀]-1', value='http://blog.naver.com/dudgus819/220942653994  [김구라의 말싸움 대행서비스] <ㅡ욕\n'
                                                'http://blog.naver.com/dudgus819/220935900355 [스타크래프트 공허의유산 1]\n'
                                                'http://blog.naver.com/dudgus819/221132921416 [오버워치- 명예와 영광을(라인하르트)]\n'
                                                'http://blog.naver.com/dudgus819/221184754402 [삐쭈코인(장삐쭈)​]\n'
                                                'http://blog.naver.com/vat19/220814826851 - 혼밥하는 만화 [1남/1녀/2공]\n'
                                                'http://blog.naver.com/vat19/220801136530 - 앵그리버드 더 무비 4 [2남/2공]\n'
                                                'http://blog.naver.com/vat19/220819235469 - 거침없이 하이킥 [2남/2녀]\n'
                                                'http://blog.naver.com/vat19/220842393433 - 2016 무한상사 1 [4남]\n'
                                                'http://blog.naver.com/vat19/220841591324 - 2016 무한상사 5 [4남]\n'
                                                'http://blog.naver.com/vat19/220837294180 - 셜록 E101 -1 [3남/1녀]\n'
                                                'http://blog.naver.com/vat19/220837363408 - 셜록 E101 -2 [3남/1녀]\n'
                                                'http://blog.naver.com/vat19/220842358644 - 주토피아 6 [2남/2녀]\n'
                                                'http://blog.naver.com/vat19/220853934141 - 갓슈벨 1-1 [2남/1녀/1공]\n'
                                                'http://blog.naver.com/vat19/220896846617 - 마이펫의 이중생활 1 [3남/1녀]\n'
                                                'http://blog.naver.com/vat19/220898867400 - 마이펫의 이중생활 3 [3남/1녀]\n'
                                                'http://blog.naver.com/vat19/220906114758 - 너에게 닿기를 15화 라이벌 [1남/3녀]',
                        inline=False)
        embed.add_field(name='[3남 1녀]-2', value='http://blog.naver.com/vat19/220919273559 - 동주 3 [3남/1녀]\n'
                                                'http://blog.naver.com/vat19/220944581076 - 12살 1-2 [1남/2녀/1공]\n'
                                                'http://blog.naver.com/vat19/220950959885 - 화이 1 [4남]\n'
                                                'http://blog.naver.com/vat19/220956802433 - 주토피아 7 [2남/2녀]\n'
                                                'http://blog.naver.com/vat19/220969398728 - 월간순정 노자키군 9화-1 [2남/2녀]',
                        inline=False)
        embed.add_field(name='[4남]', value='원펀맨 Ep.2-2 http://blog.naver.com/bbyong456/220549770966\n'
                                           '돌격!크로마티고교 Ep.6 http://blog.naver.com/bbyong456/220548812208\n'
                                           '원펀맨 Ep.3 http://blog.naver.com/bbyong456/220614257867\n'
                                           '돌격!크로마티고교 Ep.3 http://blog.naver.com/bbyong456/220060174125\n'
                                           '원펀맨 3화 http://blog.naver.com/cksdnda241/220568634123\n'
                                           '데스노트 10화 의혹 http://blog.naver.com/snowrain1100/221022379359\n'
                                           '나루토질풍전 357화 아마테라스 http://blog.naver.com/snowrain1100/221083268405',
                        inline=False)
        embed.add_field(name='[3남 1녀]', value='요괴워치 1화-1 http://blog.naver.com/teajin288/220579469824\n'
                                              '원펀맨 Ep.1-1 http://blog.naver.com/bbyong456/220539704735\n'
                                              '하루 http://blog.naver.com/ghfkddl67/220320006082\n'
                                              '나츠메우인장 0101 http://blog.naver.com/allstar0_7/50079103433\n'
                                              '마기(3) http://blog.naver.com/dpwl970419/220216529584\n'
                                              '도쿄구울(1-2) http://blog.naver.com/dpwl970419/220218607101\n'
                                              '바라카몬 (1) http://blog.naver.com/dpwl970419/220224726802\n'
                                              '페어리테일 93화-나는 여기에 서있어 http://blog.naver.com/snowrain1100/221060456830\n'
                                              'K <RETURN OF KINGS> 7화 http://blog.naver.com/snowrain1100/221052716421\n'
                                              'K <RETURN OF KINGS> 5화 http://blog.naver.com/snowrain1100/221056213378',
                        inline=False)
        embed.add_field(name='[2남 2녀]', value='고스트 스위퍼 [유령사무소 출동개시] http://blog.naver.com/osuk0070/10090967977\n'
                                              '월간소녀 노자키군 Ep.3 http://blog.naver.com/bbyong456/220614482756\n'
                                              '다다다15화(1단락) http://blog.naver.com/bbyong456/220056072582\n'
                                              '달빛천사 1화 http://blog.naver.com/carl0508/10128069854\n'
                                              '스캣댄스 3화 http://blueattack.blog.me/30137421369\n'
                                              '스캣댄스 4화 http://blueattack.blog.me/30137471050\n'
                                              '월간순정 노자키군(5-1) http://blog.naver.com/dpwl970419/220235949131\n'
                                              '월간순정 노자키군(5-2) http://blog.naver.com/dpwl970419/220306686644\n'
                                              '월간순정 노자키군 (3-2) http://blog.naver.com/dpwl970419/220215328861\n'
                                              '월간순정 노자키군(6-1) http://blog.naver.com/dpwl970419/220349883494\n'
                                              '바라카몬 (2-1) http://blog.naver.com/dpwl970419/220320775266\n'
                                              '월간순정 노자키군 http://blog.naver.com/dpwl970419/220218653984\n'
                                              '웹드라마<스타트 러브>1화 http://blog.naver.com/snowrain1100/221064846243',
                        inline=False)
        embed.add_field(name='[2남 1녀 1공]', value='A.I-1 http://stephanos_.blog.me/30145702327\n'
                                                 'A.I-2 http://stephanos_.blog.me/30145703376', inline=False)
        embed.add_field(name='[1남 3녀]', value='기동아 부탁해 Episode.1 http://blog.naver.com/bbyong456/220707431052\n'
                                              '요츠바랑 Episode.1 http://blog.naver.com/bbyong456/220707864442\n'
                                              '기동아 부탁해 Episode.2 http://blog.naver.com/bbyong456/220707898623\n'
                                              '이웃집 토토로 2부 http://blog.naver.com/ghfkddl67/220503177173\n'
                                              '달빛천사 32화 http://blog.naver.com/carl0508/10128776880\n'
                                              '아따맘마 201화-엄마주먹밥이 무서워 http://blog.naver.com/v_oice/120110503285\n'
                                              '이누야샤 1기 3화 http://blog.naver.com/v_oice/120112090757', inline=False)
        embed.add_field(name='[1남 2녀 1공]', value='짱구는못말려-엄마는 모범운전사 http://blog.naver.com/v_oice/120110621468',
                        inline=False)
        embed.add_field(name='[1남 1녀 2공]', value='A.I-3 http://stephanos_.blog.me/30145725947\n'
                                                 '검정고무신 1화 - 만우절 http://blog.naver.com/v_oice/120111964782',
                        inline=False)
        embed.add_field(name='[4녀]', value='리틀 위치 아카데미아 part.1 http://blog.naver.com/bbyong456/220300575943',
                        inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith("~인원별대본 5"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <5명>',
            description=':arrow_down:',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[2남 2녀 1공]', value='http://blog.naver.com/dudgus819/220937207590 [개구리 중사 케로로]\n'
                                                 'http://blog.naver.com/dudgus819/221058762365 [보스베이비 -2]', inline=False)
        embed.add_field(name='[3남 1녀 1공]', value='http://blog.naver.com/dudgus819/221057719910 [보스베이비 -1]\n'
                                                 'http://blog.naver.com/dudgus819/221112162556 [보스베이비 -5]', inline=False)
        embed.add_field(name='[2남 3녀]', value='http://blog.naver.com/dudgus819/220946763535  [오버워치 잠입 (솜브라)]\n'
                                              'http://blog.naver.com/dudgus819/220992004027  [너의 이름은-1]', inline=False)
        embed.add_field(name='[3남2녀]', value='http://blog.naver.com/dudgus819/220994448406   [너의 이름은-4]\n'
                                             'http://blog.naver.com/dudgus819/220995526006   [너의 이름은-6]', inline=False)
        embed.add_field(name='[4남1녀]-1', value='http://blog.naver.com/dudgus819/221144491879 [가디언즈 -2 ]\n'
                                               'http://blog.naver.com/vat19/220819969444 - 전설의 마법 쿠루쿠루 2화 [2남/2녀/1공]\n'
                                               'http://blog.naver.com/vat19/220819235469 - 거침없이 하이킥 [3남/2녀]/[2남/3녀]\n'
                                               'http://blog.naver.com/vat19/220802711926 - 앵그리버드 더 무비 1 [3남/2녀]\n'
                                               'http://blog.naver.com/vat19/220824940675 - 앵그리버드 더 무비 2 [4남/1녀]', inline=False)
        embed.add_field(name='[4남1녀]-2', value='http://blog.naver.com/vat19/220893567628 - 너에게 닿기를 3화 방과후-1 [2남/3녀]\n'
                                               'http://blog.naver.com/vat19/220895243396 - 너에게 닿기를 3화 방과후-2 [2남/3녀]\n'
                                               'http://blog.naver.com/vat19/220896846617 - 마이펫의 이중생활 1 [3남/2녀]\n'
                                               'http://blog.naver.com/vat19/220897186368 - 마이펫의 이중생활 2 [3남/2녀]\n'
                                               'http://blog.naver.com/vat19/220902365050 - 마이펫의 이중생활 4 [3남/2녀]\n'
                                               'http://blog.naver.com/vat19/220921738964 - 빅 히어로 1 [3남/1녀/1공]\n'
                                               'http://blog.naver.com/vat19/220928171463 - GTO 1-1  [4남/1녀]\n'
                                               'http://blog.naver.com/vat19/220944581054 - 12살 1-1  [1남/3녀/1공]', inline=False)
        embed.add_field(name='[5남]', value='원펀맨 Ep.2-2 http://blog.naver.com/bbyong456/220549770966\n'
                                           '돌격! 크로마티 고교 Ep.8 http://blog.naver.com/bbyong456/220706111405\n'
                                           '돌격! 크로마티 고교 Ep.9 http://blog.naver.com/bbyong456/220707066230\n'
                                           '돌격!크로마티고교 Ep.7 http://blog.naver.com/bbyong456/220562188907\n'
                                           '돌격!크로마티고교 Ep.1~2 http://blog.naver.com/bbyong456/220056125551', inline=False)
        embed.add_field(name='[4남 1녀]', value='더파이팅 챔피언 로드 http://blueattack.blog.me/30069501228\n'
                                              '달수 이야기 (1~3) http://blog.naver.com/dpwl970419/220396904502', inline=False)
        embed.add_field(name='[3남 2녀]', value='갓파쿠와 여름방학을 1부 http://blog.naver.com/ghfkddl67/220493883745\n'
                                              '명탐정 코난 193화 http://blog.naver.com/carl0508/10150675958\n'
                                              '주군의 태양 4회 http://stephanos_.blog.me/30179677611\n'
                                              '연애혁명(1~3) http://blog.naver.com/dpwl970419/220367590138\n'
                                              '마기(1-1) http://blog.naver.com/dpwl970419/220209849144\n'
                                              '일곱개의 대죄(1-1) http://blog.naver.com/dpwl970419/220314134207\n'
                                              '중2병이라도 사랑이 하고싶어1화 http://blog.naver.com/snowrain1100/221021130966', inline=False)
        embed.add_field(name='[2남 3녀]', value='귀를 기울이면 1부 http://blog.naver.com/ghfkddl67/220512150396\n'
                                              '너에게 닿기를 2화 http://blueattack.blog.me/30137545694\n'
                                              '주군의 태양 2회-2 http://stephanos_.blog.me/30177097205\n'
                                              '월간순정 노자키군 (3-1) http://blog.naver.com/dpwl970419/220214310056\n'
                                              '아따맘마 201화-예술의 가을 http://blog.naver.com/v_oice/120110542039', inline=False)
        embed.add_field(name='[2남 2녀 1공]', value='기생수 1화 http://blog.naver.com/ghfkddl67/220558111299\n'
                                                 '기생수 1화 http://blog.naver.com/ghfkddl67/220558111299\n'
                                                 '기생수 1화 http://blog.naver.com/ghfkddl67/220558111299', inline=False)
        embed.add_field(name='[2남2공1녀] ', value='강철의 연금술사 리메이크 1화 http://blog.naver.com/specialist_x/90046234671', inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith("~인원별대본 6"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <6명>',
            description=':arrow_down:',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[4남 1녀 1공] ', value='http://blog.naver.com/dudgus819/220959902147 [마다가스카의 펭귄 1화]\n'
                                                  'http://blog.naver.com/dudgus819/221116511500 [보스베이비-6]', inline=False)
        embed.add_field(name='[3남 2녀 1공]', value='http://blog.naver.com/dudgus819/221060422135 [보스베이비 -3]\n'
                                                 'http://blog.naver.com/dudgus819/221143110463 [가디언즈 -1]', inline=False)
        embed.add_field(name='[3남 3녀]', value='http://blog.naver.com/dudgus819/220992660440  [너의 이름은-2]\n'
                                              'http://blog.naver.com/dudgus819/220993376128  [너의 이름은-3]\n'
                                              'http://blog.naver.com/dudgus819/220994918554  [너의 이름은-5]', inline=False)
        embed.add_field(name='[4남 2녀]', value='http://blog.naver.com/snowrain1100/221052604993 이누야샤 완결편 15화\n'
                                              'http://blog.naver.com/snowrain1100/221156302734 일곱개의대죄 외전\n'
                                              '헌터X헌터 Ep.1 http://blog.naver.com/bbyong456/220544651130\n'
                                              '캐치미이프유캔 http://blog.naver.com/bbyong456/220548659030\n'
                                              '어벤져스 Part.1 http://blog.naver.com/bbyong456/220701935783\n'
                                              '갓파쿠와 여름방학을 2부 http://blog.naver.com/ghfkddl67/220498661023\n'
                                              '타이밍 중-1 http://blog.naver.com/teajin288/220580413279\n'
                                              '성검의 블랙스미스 0101 http://blog.naver.com/allstar0_7/50078843621', inline=False)
        embed.add_field(name='[5남 1녀]', value='A특공대 Ep.2 http://blog.naver.com/bbyong456/220626859257\n'
                                              '원피스 극장판 9기 http://blog.naver.com/ghfkddl67/220334615037', inline=False)
        embed.add_field(name='[5남 1공]', value='돌격!크로마티고교 Ep.5 http://blog.naver.com/bbyong456/220061968066', inline=False)
        embed.add_field(name='[4남 1녀 1공]', value='진격의 거인 http://blog.naver.com/ghfkddl67/220360918992', inline=False)
        embed.add_field(name='[3남 3녀]', value='늑대아이 1부 http://blog.naver.com/ghfkddl67/220322749509\n'
                                              '다다다15화(2단락) http://blog.naver.com/bbyong456/220056115462\n'
                                              '고양이의 보은 1부 http://blog.naver.com/ghfkddl67/220531887261', inline=False)
        embed.add_field(name='[1공 3남 2녀]', value='디지몬 어드벤처 47화 http://blog.naver.com/carl0508/10122442150', inline=False)
        embed.add_field(name='[3남 2녀 1공]', value='기생수 1화 http://blog.naver.com/ghfkddl67/220558111299', inline=False)
        embed.add_field(name='[3남 3공]', value='헌터X헌터 Ep.2 http://blog.naver.com/bbyong456/220563510483', inline=False)
        embed.add_field(name='[2남 4녀]', value='달빛천사 9화 http://blog.naver.com/carl0508/10127705351', inline=False)
        embed.add_field(name='[2남 2녀 2공]', value='빨간망토 차차 Ep.1 http://blog.naver.com/bbyong456/220545316593', inline=False)
        embed.add_field(name='[1남 4녀 1공]', value='다다다 63화 http://blog.naver.com/peachyolk/220280328549\n'
                                                 'http://blog.naver.com/vat19/220799344400 - 명탐정코난 순흑의 악몽 2 [3남/3녀]\n'
                                                 'http://blog.naver.com/vat19/220819235469 - 거침없이 하이킥 [3남/3녀]\n'
                                                 'http://blog.naver.com/vat19/220822803262 - 주토피아 4 [4남/2녀]\n'
                                                 'http://blog.naver.com/vat19/220825593311 - 라이언 킹 1 [2남/2녀/2공]\n'
                                                 'http://blog.naver.com/vat19/220877376336 - 보루토 - 나루토의 취임식 [2남/3녀/1공]\n'
                                                 'http://blog.naver.com/vat19/220881558179 - 너에게 닿기를 2화 자리바꾸기 [3남/3녀]\n'
                                                 'http://blog.naver.com/vat19/220897186368 - 마이펫의 이중생활 2 [4남/2녀]\n'
                                                 'http://blog.naver.com/vat19/220853936589 - 갓슈벨 1-2 [3남/2녀/1공]\n'
                                                 'http://blog.naver.com/vat19/220922085210 - 빅 히어로 2 [3남/2녀/1공]\n'
                                                 'http://blog.naver.com/vat19/220928171463 - GTO 1-1 [4남/2녀]\n'
                                                 'http://blog.naver.com/vat19/220951835681 - 모브 사이코 100 2화 [3남/2녀/1공]', inline=False)
        embed.add_field(name='[3남3녀]', value='http://blog.naver.com/snowrain1100/221021005101 괴도키드1412 14화 크리스탈 마더\n'
                                             '​http://blog.naver.com/snowrain1100/221021783293 나만이 없는 거리 1화 주마등\n'
                                             'http://blog.naver.com/snowrain1100/221065655130 장애인식개선 웹드라마 아름다운소통\n'
                                             'http://blog.naver.com/snowrain1100/221153318236 클로저스 1화 시작 검은양 팀\n'
                                             'http://blog.naver.com/snowrain1100/221157811969 삼성화재 드라마 비바앙상블 1-2\n'
                                             'http://blog.naver.com/allstar0_7/50078210979 엑스 0101', inline=False)
        embed.add_field(name='[6남]', value='http://blog.naver.com/snowrain1100/221154532972 하이큐 1기 21화 선배의 실력', inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith("~인원별대본 7"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <7명>',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[7남]', value='A특공대 Ep.1 http://blog.naver.com/bbyong456/220626310164', inline=False)
        embed.add_field(name='[6남 1녀]', value='어벤져스 Part.2 http://blog.naver.com/bbyong456/220705022041\n'
                                              '사무라이 참프루 Ep.1 http://blog.naver.com/bbyong456/220314110281', inline=False)
        embed.add_field(name='[5남 2녀]', value='어벤져스 Part.1 http://blog.naver.com/bbyong456/220701935783\n'
                                              '암살교실 1화 http://blog.naver.com/ghfkddl67/220318950394\n'
                                              '코난 극장판 http://blog.naver.com/ghfkddl67/220402508016', inline=False)
        embed.add_field(name='[4남 3녀]', value='고양이의 보은 2부 http://blog.naver.com/ghfkddl67/220531892222\n'
                                              '코난 극장판 http://blog.naver.com/ghfkddl67/220402508016\n'
                                              '갓파쿠와 여름방학을 3부 http://blog.naver.com/ghfkddl67/220499359341\n'
                                              '연애혁명(11~12) http://blog.naver.com/dpwl970419/220399189907', inline=False)
        embed.add_field(name='[4남 3공]', value='헌터X헌터 Ep.2 http://blog.naver.com/bbyong456/220563510483', inline=False)
        embed.add_field(name='[4남 2녀 1공]', value='헌터X헌터 Ep.1 http://blog.naver.com/bbyong456/220544651130', inline=False)
        embed.add_field(name='[4남 1녀 2공]', value='헌터X헌터 Ep.3 http://blog.naver.com/bbyong456/220564316878', inline=False)
        embed.add_field(name='[3남 4녀]', value='주군의 태양 3회-2 http://stephanos_.blog.me/30177319522\n'
                                              '올림포스 가디언 4화 http://blog.naver.com/ghfkddl67/220368122145', inline=False)
        embed.add_field(name='[2공 3남 2녀]', value='디지몬 어드벤처 20화 http://blog.naver.com/carl0508/10126804533', inline=False)
        embed.add_field(name='[1남/3녀/3공]', value='명탐정코난 순흑의 악몽 http://blog.naver.com/vat19/220797935909', inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith("~인원별대본 8"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <8명>',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[5남 3녀]', value='토이스토리3 http://blog.naver.com/ghfkddl67/220531904217\n'
                                              '주토피아 4 http://blog.naver.com/vat19/220822803262', inline=False)
        embed.add_field(name='[6남 2녀]', value='페어리테일 세븐드래곤즈 http://blog.naver.com/cksdnda241/220666805872', inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith("~인원별대본 9"):
        channel = message.channel
        embed = discord.Embed(
            title='인원별 대본 목록 <9명>',
            colour=discord.Colour.gold()
        )

        embed.set_footer(text='END')

        embed.add_field(name='[6남 3녀]', value='원피스,토리코,드래곤볼 합작 http://blog.naver.com/ghfkddl67/220325076692', inline=False)
        embed.add_field(name='[3남 4녀 2공]', value='http://blog.naver.com/snowrain1100/221152426605 엘소드:엘의여인 6화', inline=False)

        await client.send_message(channel, embed=embed)


    if message.content.startswith('~단문'):
        shortscript = [
            "왜.. 죽였냐고? 3개월 전에 모녀 살인사건 있지..? 그거... 내가 한거야... 근데.... (헛웃음) 그걸 딴 새x가 했다고 하네? 야... 네가 공적을 하나 세웠어.. 근데 그걸 딴 새x가 했다고 해봐... "
            "기분이 나빠 안 나빠.. 시x 기분 더럽잖아.. 그래서 죽였어...오늘 자수하러 온 것도.... 이거 말하려고 온거야.",
            "내가 그렇게 못나고 바보 같고 이상한가요? 왜 친구들은 날 싫어하죠? 친구들에게 반갑게 다가가면 날 자꾸만 피해요. 제발요. 그 이유를 알려주면 안 되나요? 뭐든지 친구들이 싫어하는 거라면 평생 안 할 자신 있어요. "
            "(울면서) 나 사실 정말, 정말 미치듯이 외롭거든요.... 나 원래부터 이런 애는 아니었는데... 무슨 방법이라도 말해주세요",
            "야! 너... 세상이 그렇게 말랑말랑한 줄 알아? 혼자서는 죽어도 살수 없는 게 세상이야. 내 얘기 잘 들어. 그 친구들, 당장은 죽기보다 싫어도 이 세상을 살아가려면 꼭 필요한 친구들이야. 너 말이야. 바보 같이 뒤로 숨지만 말고. "
            "나를 봐. 내 눈을 똑바로 보라구! 앞으로 힘들어 쓰러지는 한이 있더라도, 남들보다 한 발 더 나아가. 당당하게. 그리고 그 친구들 찾아가서 말 해. 난 바보가 아니라고. 누구보다도 용기 있고, 그 누구보다도 강하다고. "
            "너희들과 친해지고 싶다고. 내 말 알아들었지? 앞으로 절대로 우는 모습 보이지마!",
            "20대 남성 (열혈 바보) 그렇게 나오겠다 이거야?! / 너 이 자식, 각오는 돼 있겠지 간다! 간다! 간다! 간다! 간다! 간다! 간다! 간다! / 너 오늘 죽을 줄 알아! 그래, 어디 누가 이기나 끝까지 해보자! 맞을때까지 쏜다! "
            "저 미꾸라지 같은 놈! 야 기분 나쁘니까 웃지 마! 웃지 말랬지! 그렇게 웃지 말란 말이야!/ 받아라!! 받아라! 받아라! 에잇! 에잇! 에잇! 저 자식이 진짜! 그게 니 녀석의 마지막 만찬인 줄 알아~!! "
            "(공격 호 계속) 너 내 손에 잡히면 죽을 줄 알아! 니가 감히 내 푸딩에 손을 대? 받아라 에잇! 에잇! 에잇! / 내 푸딩의~ / 복수다!!!",
            "KBS 무대 <수호천사119> 中  (N) 일하는 날이면 어김없이 시체를 본다. 드물 땐 한 구, 많을 땐 하루에 세 구까지도. 소멸하는 수많은 인명을 보며, 타인의 죽음에 무감각해지고, 일에 대한 죄책감도 무뎌져갔다. "
            "진짜 소방대원들은, 나처럼 수도 없이 마주할 그 생과 사의 교차점에서, 정말 간절하게 일하고 있을까? 남의 목숨을 살리기 위해 정말로 진심을 다한다는 일이 가능한 걸까? 목숨을 던질 만큼...? 나는... 나도... 그럴 수 있을까.?",
            "엄마. 엄마 딸 보은이 여기 봐봐. 나 대법관 됐어. 엄마 기쁘지? 나 멋있어? 이 옷 잘 어울리지? 거봐. 나 됐잖아. 엄마 근데....근데 엄마 나 이제 못 만나. 왜냐면 나 대법관 되서 하늘나라 가. "
            "거기서 나쁜 놈들 심판해 줄꺼야 내가. 그러니까 엄마 나 만난다고 직장 찾아오고 그러면 안돼. 응? 엄마 알았지? 엄마? 알았어? 그리고 있잖아... 엄마 우리 옛날에 엄마 아파가지고 병원에 누워있을 때, "
            "그때 나 못 먹어가지고 영양실조 걸리고 엄마 아파가지고 깨어나지도 못하고 그 때 기억나지? 그 때 우리 참 힘들었다. 그치? 그래가지고 오빠가 엄마랑 나랑 둘 다 죽일라고... 근데 엄마 이거 다 알았다면서... "
            "어? 엄마 다 알고 있었다면서...맞지? 그래서 오빠가, 이 병신 같은 게...엄마한테 미안해가지고 다시는 집에도 못 오고 죽어라고 일해가지고 돈 다 준거잖아. 엄마, 자홍이... 자홍이 그 병신 같은게 평생을 그랬다고... 엄마.",
            "엄마. 엄마 아들 수홍이. 여기 봐봐. 나 대법관 됐어. 엄마 기쁘지? 나 멋있어? 이 옷 잘 어울리지? 거봐. 나 됐잖아. 엄마 근데....근데 엄마 나 이제 못 만나. 왜냐면 나 대법관 되서 하늘나라 가. "
            "거기서 나쁜 놈들 심판해 줄꺼야 내가. 그러니까 엄마 나 만난다고 부대 찾아오고 그러면 안돼. 응? 엄마 알았지? 엄마? 알았어? 그리고 있잖아... 엄마 우리 옛날에 엄마 아파가지고 병원에 누워있을 때. "
            "그때 나 못 먹어가지고 영양실조 걸리고 엄마 아파가지고 깨어나지도 못하고 그 때 기억나지? 그 때 우리 참 힘들었다. 그치? 그래가지고 형이, 자홍이 이 새끼가 엄마랑 나랑 둘 다 죽일라고... "
            "근데 엄마 이거 다 알았다면서... 어? 엄마 다 알고 있었다면서...맞지? 그래서 형 이 새끼가, 이 병신 같은 게...엄마한테 미안해가지고 다시는 집에도 못 오고 죽어라고 일해가지고 돈 다 준거잖아. "
            "엄마, 자홍이... 자홍이 그 병신 같은게 평생을 그랬다고... 엄마.",
            "그거 인정하면, 뭐가 달라지는데요? 제가 누차 말씀드렸죠, 경위님이 곱씹어야 할 건 내 신상같은게 아니라 당신이 지은 죄에요. 미연이 죽음을 방관한 것도 모자라서 푼돈에 팔아넘긴 파렴치한 죄! "
            "마찬가지로 지금 그 물탱크에서 경위님을 구해줄 수 있는 것도 밖에 있는 동료가 아니라,  마음에서 우러난 후회와 반성이라고, 내가 도대체 몇번을 말해줘야 알아 쳐먹겠냐구요?! 예?!",
            "우리 엄마.. 우릴 얼마나 끔직히 생각했는데.. 수진이가 네 살 땐가 다섯 살 땐가.. 갑자기 한밤중에 토하구 열이 올라 펄펄 끓은 적이 있었어. 그 때 수진이 업고 엄마랑 나랑 밤길을 달려서 병원에 갔는데.. "
            "우리 엄마 발이 어떻게 됐는지 알아? 급한 김에 슬리퍼를 신고 달렸는데 눈길에 다 벗겨진 거야. 자갈에 찢겨서 피가 나고 발이 새빨갛게 얼었는데두 울 엄마 수진이 걱정만 했어. 그런 게 엄마야 임마. "
            "세상에 물에 빠지면 저 혼자 살겠다고 몸부림치는 부모가 어딨냐? 자기 팔자 고친다고 자식 버리는 엄마가 어딨어?",
            "조윤주 정신 차려, 우린 이미 끝난 사이야. 나는 곧 아이 아빠가 된다고 그만, 그만해 너는 미쳤어. 이제 제발 정신 좀 차리라고. 그날도 너희 집에 날 부른 날도, 너 계획적으로 일부러 불렀던 거지? "
            "아영이 불러 놓고는, 날 협박해서 일부러 너희 집에 오게 했던 거잖아. 그래서 뭘 어쩌겠다고? 네가 원하는 게 뭔데? 아니... 그래도 나는 너에게 안 돌아가. 만약에... 아영이가 이혼 하겠다면.. "
            "그래 나는 죄인이니까... 아영이가 원하는 데로 해줄 수밖에 없어,하지만 나는 평생 속죄 하면서, 어떻게든 아영이 옆에서 아영이와 내 아이 지킬 거야... 빌고 또 빌면서 평생을 아영이가 용서해 줄때까지 아영이와 내 아이 옆에 맴돌 거라고. "
            "너는 내 실수일 뿐이야",
            "네 맞아요. 나는 마녀에요. 마법의 힘으로 누군가의 소원을 이뤄주는 일을 하죠. 마법을 부리려면 적당한 장소가 필요해요. 아무 곳에서나 할 수는 없죠. 이 식당 자리가 딱이에요. "
            "야트막한 언덕 위, 사거리에 있으면서 빛이 잘 들어오지도 않고 무엇보다 사람들 눈에 잘 띄지 않는 곳. 그러면서도 너무 외지지 않는 그런 곳! 얼마나 찾아 헤맸는지 몰라요. "
            "얼마나 이 날을 기다렸는지...이 식당을 저에게 주세요.",
            "새 식구가 오던 주말, 백은호와 백아영은 결혼식장에 나란히 앉아 있었다. 앉아 있으면 둘은 뒷모습이 닮아 보였다. 양쪽 귀에서 목을 지나 어깨에 이르는 선이 왠지 모르게 비슷했다. "
            "‘피가 전혀 안 섞인 남남’은 아닐지도 모른다는 생각이 들 만큼, 둘은 딱 그만큼 비슷했다. 백은호 오른편에 앉은 강윤희는 피곤한 표정으로 눈을 감고 있었다. 강윤희는 아버지의 또 다른 남동생 가족이 모여 있는 테이블로 시선을 돌렸다. "
            "강윤희의 아버지는 남동생이 둘이었다. 강윤희의 형제들은 아버지 바로 밑의 남동생을 큰 작은아버지, 아버지와 터울이 많이 진 막내 남동생을 작은 작은아버지라고 불렀다. "
            "강윤희는 결혼식장에 도착하면서부터 작은 작은아버지 옆에 붙어 있는 한 소년을 보고 있었다. 소년은 식장에 도착한 강윤희에게 ‘안녕하세요 누나’ 그렇게 말하며 다가왔다. 강윤희는 당황스러웠다. "
            "소년은 강윤희의 아버지의 막내 남동생의 아들이었고, 강윤희의 기억 속에선 병약한 남자아이일 뿐이었다. 몇 년 사이에 이렇게 성인 남자에 가까운 모습을 하고 있을 줄 알았다면 강윤희는 작은 작은어머니의 부탁을 거절했을 것이다. "
            "식장을 나서기 전, 강윤희 앞으로 작은 작은아버지가 다가왔다. 그가 강윤희의 손을 모아 잡았다. 아버지의 아주 어린 동생이었던 남자. 강윤희는 강중식을 무표정하게 바라보았다. "
            "사포처럼 거친 강중식의 가운뎃손가락이 강윤희의 손등을 눌렀다. 강윤희는 슬그머니 손을 뺐다. 곧 쉰이 되는 강중식은 열 살 이상 터울인 형들과 비슷한 연배로 보일 정도로 급속히 늙어가고 있었다. "
            "차려입은 양복만 아니라면 세탁기에서 탈수되다 나온 것처럼 엉망인 모습이었다. 강윤희는 강중식 옆에 선 소년을 착잡한 마음으로 바라보았다.",
            "단조로운 일상만 반복될 것 같은 수영장에서도 가끔은 이상한 일이 벌어진다. 언젠가 어떤 아주머니는 수영모자와 물안경만 쓰고 샤워장을 나섰다. "
            "샤워하느라 벗어놓은 수영복은 그대로 샤워기 조절 레버 위에 걸어둔 채였다. 이미 풀에 들어와 있던 다른 여자들이 어서 돌아가라고 열렬히 손짓했지만 그녀는 그걸 어서 오라는 신호로 받아들인 것 같았다. "
            "더 잰 걸음으로 풀을 향해 달려왔다. 그러고는 풀 앞에서 잠시 멈칫하더니 풍덩 물속으로 뛰어들었다. 수영장에는 잠시 침묵이 흘렀다. 여자 강습생들이 물속에 갇힌 그녀를 둥그렇게 에워쌌다. "
            "검은색 아레나 수영복을 입은 여자가 사다리를 잡고 풀 밖으로 나와 샤워장으로 뛰어갔다. 수영장의 모든 눈이 그 여자의 뒷모습을 주시하고 있었다. 잠시 후, 그녀는 손에 수영복을 들고 다시 나타났다. "
            "자기가 벌거벗기라도 한 것처럼 여자는 위축되어 있었다. 곧 그녀의 손으로부터 수영복이 전해졌다. 수영복의 주인은 울 것 같은 표정으로 조심스레 수영복을 꿰어 입었고 한 동료가 그녀의 수영 모자를 벗겨 자기 것과 바꾸어 썼다. "
            "강습이 재개되었다. 강습생들이 일제히 팔을 젓고 발을 차며 차례차례 앞으로 나아갔다. 다들 비슷비슷한 모습이어서 누가 조금 전 소동의 주인공인지 금세 알 수 없게 되었다. "
            "그렇지만 아무것도 모른 채 태연하게 풀을 향해 걸어 나오던 그 아주머니의 벌거벗은 모습이 뇌리에서 쉽게 사라지지 않았다. 그것은 슬프다고도, 그렇다고 우습다고도 할 수 없는 기묘한 이미지였다.",
            "(10대 여) 모두 다 알다시피, 우리 걸스카우트의 가장 작은 단위를 ‘보’라고 해. 보통 다섯 명씩 모여서 한 보를 이루고 그 보를 통솔하는 사람을 보장이라고 해. "
            "보장은 당연히 6학년인 내가 맡고 부보장은 그 다음으로 높은 학년이 맡는 거야. 다들 알지? 근데... 5학년 이다빈- 넌 걸스카우트란 애가 대답도 크게 못해? "
            "이다빈- 넌 우리 보의 부보장이란 게 여태 걸스카우트 선서도 못 외웠어. 다른 동생들한테 창피하지도 않아? 나중에 내가 시켜볼 거야. 아 참, 그리고 다음 주 야영 갈 때 버너랑 쌀도 니가 갖고 와. "
            "시키는 대로 안 해? 난 보장이니까 할 게 많단 말이야.",
            "(30대 여, 약간 미친) 양춘덕 씨를 아세요? 잔혹낙원의 저자를 아시냐구요... 전 양춘덕 선생님의 문하생이었어요. 선생님의 작품이라면 누구보다 잘 알아요. "
            "잔혹낙원은 선생님 거예요, 명백한 증거가 제게 있으니 확인 시켜드릴 수도 있어요, 이 많은 사람들 앞에서 당장... 전 강재현 씰 이해한다구요. 선생님은 돌아가셨고 작품은 남았으니... "
            "제가 어떻게 해드리길 바라세요? 입 다물어드림 될까요? 그래요, 이 상황에 진실이 밝혀지면 강재현 씬 난처한 상황에 빠지게 될 거예요. 도피도 불사해야 할 만큼... 그렇죠? 입 다물어 드릴게요! "
            "그럼 소설 수익의 절반만 주세요. 그럼 저도 공범이 되니 입을 다물 수밖에 없을 테고 그렇게 되면 강재현 씨도 절 믿으실 수 있으실 거예요.",
            "(늙은 주술사) 아야후아스카~ 흐흐...아야후아스카는 포도넝쿨처럼 생겼어. 아마존의 식물이야. 녹색 잎을 우려내 마시면 강력한 환각에 빠지게 되지. 하지만 부작용은 없어. "
            "중독도 되지 않고 금단현상도 없지. 아마존 사람들은 아야후아스카를 단순한 환각제로 분류하는 걸 거부해. 우리 아마존 샤먼(shaman)들은 병자가 찾아오면 아야후아스카를 복용하고 식물의 영혼과 접촉해 처방법을 얻어내지. "
            "우리한테 아야후아스카야는 신성한 식물이야. 페루 ‘파블로 아마링고’라는 사람이 아야후아스카를 먹고 그림을 그렸어. 그 그림 얼만지 알아? 현재 한 점에 8천 달러 이상에 거래되고 있지. 으흐흐흐.. 아야후아스카~",
            "(10대 여) 오전 아홉시 이십분의 빛을 놓치면 안 되었다. 아홉시 이십분에서 오십분까지의 빛은 형태의 가장자리를 넓고 투명하게 만드는데, 서서히 엷어지다가 투명해지는 그 지점을 자연스럽게 그려내는 것이 내 목표였다. "
            "커튼 틈으로 들어온 직선의 빛이 선생님의 머릿결과 귓불, 어깨와 팔에 부딪혀 곡선으로 튕겨 나가는 장면을 기억해 두었다. 가르마에서 반사되는 빛은 아주 투명하지만 머리칼의 경계 때문에 그리는 것에 한계가 있었다. "
            "가장 눈부시고 투명한 빛은 불룩한 옆구리에서 반사되는 빛이었는데 그게 참 안타까웠다.",
            "상주실 한쪽에 조그맣게 쪼그리고 앉아있는 여인은 상복도 입지 않고 있었다. 주름이 더는 잡힐 데가 없을 지경으로 쪼글쪼글한 얼굴, 그 얼굴은 지난 날 겪어온 고생이 얼마나 극심했는지를 비춰주는 거울이었다. "
            "여자가 겪어낸 간난신고는 얼굴하고 손에 그려진다는 옛말대로, 시집와서 꽃피는 날을 한 번도 보지 못했을 여인이 한없이 가련하고 애처로워 윤 혁은 가슴이 아팠다. "
            "아무 느낌도 담기지 않은 목소리로 대꾸한 여인은 아들에게 눈길을 보내는가 싶더니, 고개를 곧 떨구어버렸다.",
            "(10대 남) 제가 죽인 거예요. 제가... 여름 캠프에 간 날, 동생이 죽었어요. 겨우..하룻밤 자고 온 건데.. 집에 돌아와서..동생을 찾는데 없어서, 어디 갔냐고 물었는데 아빠한테 갔다고 해서.... "
            "근데 신발도 그대로 있고, 동생이 없으면 울고불고 난리치는 토끼인형도 그대로 있고... 그래서, 이상해서... 온 집안을 다 뒤졌어요. 시끄럽게 굴면 그 년이 또 미친 짓할 테니까 까치발을 들고.. "
            "‘희섭아..’ 그렇게 방안 벽장을 열어보니, 동생이 있었어요. 머리를 이렇게, 푹 숙이고...얼른 안아서 깨웠는데.. 차가운 거예요. 이상해서 얼굴을 보니까...입이 이상해...빨개..온통. 그래서... 그냥 도망쳤어요.",
            "(30대 남) 가짜를 가짜라고 밝히는 게 뭐, 어때서. 여자친구인척, 친구인척, 환자인척 진짜도 아니면서 가짜 노릇하는 것들은 한번 혼쭐이 나봐야 정신 차리지. 다시는 가짜 노릇을 못하게 해야 한다고! "
            "가짜들이 판을 치게 되면 진짜들이 가짜로 오해받기 쉽다고. 이 사회가 가짜들로 득실 되면 좋겠어? 그렇지! 그러니 사회 정의를 구현하기 위해서 우리가 힘을 합쳐 가짜들을 물리쳐보자 이거야. "
            "지금 이 시간부터 우린 이 방을 나가서 다시는 돌아오지 않는 거야. 물론 내일 검진시간이 변경된 것도 그들에게 알려선 안 되겠지. 좋았어! 지금부터 우린 같은 배를 탄 거야. 설령 우리 배가 은조각배일지라도, 우리가 "
            "남기고 간 물살에 물결은 출렁이게 되어있다고.",
            "(늙은 뱀) 오래 전에, 어느 화물선박 방역책임자가 골칫거리 쥐떼들을 박멸한다고, 배에다 청산가스를 잔뜩 뿌렸다네. 골칫거리들이 몰살당했을까? 천만에. 죽음의 가스가 들어오는 파이프를 봉쇄하는 계략을 썼지. "
            "누가 희생을 했겠나. 나이가 가장 많은 늙은 쥐들이, 몸뚱이로 파이프 구멍을 겹겹이 틀어막고, 종족을 살렸네. 콘크리트나 철근처럼 단단한 것을 갉아서 부서뜨리고, 썩은 물이나 바닷물로도 우린 갈증을 해소할 수 있어. "
            "그뿐인가? 영하 40도, 영상 60도의 악조건에서도 우린 거뜬히 생존해....자네도 이를 악물고 우리처럼 무소불위 악전고투했지. 어떤가. 이제 나랑 허물없이 교감할 수 있겠나?",
            "(30대 남) 인류의 종말 예고가 발표되면 으레 있을 것이라고 믿었던 폭동이나 약탈은 영화에서만 있는 일인 듯 서울은 지나치게 고요했다. 각종 종교단체의 교당은 예상외로 조용했고, 사람들은 오래 생각해둔 여행을 준비하듯 침착했다. "
            "종말을 앞둔 지구는 숨 막히게 조용했다. 나는 사람들이 믿기지 않아서보다 용기가 없어서라고 생각했다. 방송사에선 헬리콥터를 띄워 전국을 비췄다. 고속도로의 정체가 심했고 개중에는 차를 버리고 걸어가는 사람도 있었다. "
            "열심히 마당을 파고 있는 가족도 보였고 옥상에 커다란 태극기를 펼쳐 놓은 집도 보였다. 눈은 하염없이 내렸다. 모두 저마다 분주했지만 소란스럽지는 않았다. 나는 사람들이 용기가 없어서가 아니라 믿기지 않아서였다고 생각을 고쳤다.",
            "정윤은 손을 들어 인사를 한 다음 계산대에서 카드로 술값을 계산했다. 정윤은 밖으로 나서기 전 규호를 보았다. 규호는 여전히 멍한 눈빛으로 자신의 앞만 바라보고 있었다. 정윤은 입을 굳게 다물고 좌우로 고개를 흔들며 밖으로 나갔다. "
            "규호는 정윤이 앉아 있던 자리의 커피 잔을 옆으로 치우고, 거기에 소주잔을 놓았다. 규호는 혼자 술을 마실 때면 늘 그러곤 했다. 거기 누가 있기라도 한 것처럼, 누가 있었으면 좋겠다는 마음으로, 그러곤 했다. "
            "규호는 소주를 탄 생맥주를 마셨다. 의자의 천을 계속 보았다. 계속 보니 거기 누가 앉아있는 것 같기도 했다. 어디선가 바람이 불어와서 땅콩 껍질이 허공에 날렸다. 자신의 몸도 공중으로 붕 떠오르는 것 같았다. "
            "규호는 양손으로 맥주잔을 꼭 쥐었다.",
            "(20대 여) 새벽에 너 나가는 거 봤어. 마루에 앉아 있는데, 정말 막막하더라. 의지할 데가 없어지면 돌아갈 마음이 날 줄 알았는데 그게 아닌 거야. 혹시 니가 마음에 걸려서 돌아와 주지 않을까, 하릴없이 문만 쳐다보게 되더라고. "
            "니가 정말로 돌아오니까 얼마나 기뻤는지, 기뻐하는 내 자신은 또 얼마나 처량하고 한심한지. 그런데도 여전히 마음을 결정할 수가 없었어.",
            "(30대 여)(지지 않고) 얼마나 대단한 가문인데요, 어머니? 자기 마음에 들지 않는다고 배꼽도 떨어지지 않은 핏덩이를 쓰레기처럼 몰래 갖다버리는 집안 말씀이신가요? 하늘 무서운 짓을 하시고도, 밤에 잠이 오던가요? "
            "(차갑게) 인정하시든 아니든 어머니 핏줄이에요! 하늘이 두 쪽이 나도 그 사실은 바뀌지 않아요. 그리고 소리 좀 그만치세요, 아기가 놀라거든요. 저 먼저 올라가겠어요. 안녕히 주무세요.……",
            "(중년 여) 휴우……. 하마터면 나까지 쫓겨날 뻔 했네. 우왕좌왕 하는 틈에 의사 가운 주워 입고 입원실까지 들어 왔을 줄은 상상도 못하겠지? "
            "(살금살금 다가가며) 민규 씨 깨면 안 되니까 조심조심……. 어머나, 자는 모습도 어쩜 이리 천사 같을까. 퍼펙트 그 자체네. (속삭이는) 민규 씨, 내가 왔어요. 이제야 우리 둘만의 시간을 보내게 됐네요. "
            "민규 씨, 생일 축하해요…. 축하의 의미로 볼에 뽀뽀 한번만 해도 될까요?",
            "어떤 연구단체의 설문 조사 결과에서 여자들이 하루 중 행복하다고 느끼는 시간은 2시간 42분뿐이라고 한다. 하루에 단 20초도 행복하지 않았을 때 내 몸 안의 작은 생명이 신호를 보내왔다. "
            "캄캄했던 내 몸 속에 반딧불처럼 작은 불빛들이 반짝반짝 불을 밝힌 것만 같았다..24시간이 통째로 행복할 수도 있다는 사실을 그 때 처음으로 알았다……. 그러나… 불이 꺼져버렸다……. 이제 다시 암흑이다.",
            "남자는 허둥대고 있었다. 삶에 대해서. 생활에 대하여. 남자는 자신이 최근에 뭔가를 잘 잊어버린다는 것이 속상하다. 뭔가를 잊어버렸던 사건들이 오히려 더 잘 기억되고 있다는 사실이 속상하다. "
            "기억하고 싶지 않은 것들은 뚜렷했고 기억해야 할 것들은 흐릿했다. 그건 인생의 법칙이기도 했다. 삶이란 의도했든 의도하지 않았든 배신을 준비하니까. 도저히 이룰 수 없는 것들을 하는 수없이 감당하게 하는 것, 그것이 삶의 책략이다.",
            "(10대 남) 말만 사납게? 가방을 휘두르면서 두드려 패더라. 난 억울하다고, 아이스께끼의 범인이 아니라고, 그렇게 말했는데도 듣지도 않고. 갑자기 울컥하네. "
            "걔가 머리가 좋잖아. 자길 건드리는 놈은 이렇게 된다고, 시범케이스로 날 잡고 경고한 거지. 오죽하면 애들이 걔 전학 온 지 두 달 만에 딱 결론 내렸잖아. ‘성깔은 개의 경지, 성적은 신의 경지.’",
            "(30대 남) 어릴 때부터 저는 비오는 날을 좋아했어요. 아니, 우산에 떨어지는 빗소리를 듣는 걸 좋아했다는 게 맞겠네요. 그래서 범인은 내게서 가장 좋아하는 날을 빼앗고 싶었던 거겠죠. "
            "최대한 나를 고통스럽게 만든 다음에 내가 가진 모든 걸 빼앗으려고 하는 게 분명해요. 빗소리 다음엔...돈을 빼앗고... 그 다음엔 내 목숨이겠죠. 아니, 목숨 다음에 돈이려나?",
            "(중년 남) 생각해보면 나는 벽 하나를 사이에 두고 극단적인 두 가지 삶을 산 셈이라네. 꿈을 꾸는 몽상가와 움직이는 행동가, 대본을 쓰는 극작가와 무대에서 연기하는 배우. "
            "어느 쪽의 삶에 더 만족했을 것 같나? (심드렁하게) 사실 둘 다 별로였어. 너무나 오랜 세월을 좁은 감방에서 탈옥에만 몰두한 탓에 내 안의 무언가 까맣게 타버린 거야. 생에 떨림을 주는 무언가가….",
            "(남 독백) 어릴 때부터 야구를 정말 좋아하긴 했다. 민석의 장모, 그러니까 송아름의 모친이 내놓은 사진들 속에 송아름의 몸은 군데군데 멍이 시퍼렇게 들어 있었다. "
            "미처 이불로 보호하지 못한 곳을 맞은 듯한 자국이었다. 손가락... 등.. 엉덩이.. 하루 종일 그 사진을 보고 또 봤다. "
            "죽은 송아름 모친이 하는 말과 사건 파일을 읽으면서 나는 자꾸만 초등학교 동창인 혜미가 했던 말들이 머릿속을 떠나지 않았다.",
            "10대 발랄하고 도도한 소녀 - 으휴~ 대학은, 지루하고 못생기고 심각한 사람들만 가는 곳이야~ 레이첼이나 앤드류처럼 말이야! (웃음)난 백화점에 가서 쇼핑을 할 거야! "
            "패트릭이 오늘 나에게 프로포즈를 할 것 같앙~. 잘생긴 외모, 뛰어난 패션센스, 게다가 빵빵한 집안까지! 완벽한 내 남친 패트릭! 근데, 정말 결혼하자고 하면 뭐라고 하지?　"
            "음.. 도도하게 굴어야지~ 난 절대로 쉬운 여자가 아니니깐! 한번은 튕겨야지~ 흥!",
            "10대 청순하고 내성적인 고등학생 - 소스케.. 나... 비록 코코미보다 아는 것도 별로 없고, 어설프고, 서투른 거 투성 일지는 모르겠지만 그래도.. 할 수 있는 데까지 해보고 싶어 힘을 합쳐서, 모두가 한 가지를 위해 땀 흘리는 거... "
            "그게 무엇보다도 멋있는 거 같아, 츠요시처럼 말이야. 나는... 그런 게... 왠지... 그냥 좋아 (미소)",
            "20~30대 중성적이며 카리스마 있는 여성 - 그 쪽이 잘 알거에요 날 돈 있는 사람은 진심으로 상대하는 게 아니에요. 돈으로 상대하는 거지 우리 조카, 참 말 안 듣게 생겼네, 방금 잘 봤니? "
            "방금 니가 본 게 앞으로 니가 나올 세상이고, 돈 없는 사람이 공부를 해야 하는 이유야. 알아들어. 조카?",
            "모험심 강한 10대 남아 - 어른들이 우리나이 때 보물을 찾으러 갔지! '애꾸눈 윌리'인가 하는 해적의 전설이야. 그 시대 유명한 해적이야 아빠한테 들었어. 윌리는 보물을 가득 훔쳤어 루비, 에메랄드... "
            "그리고 보물을 자기 배에 싣고 멀리 어디론가 떠났대. 근데 영국 왕이 알아내서 함대로 추격을 한 거야몇 주일 뒤에 함대가 윌리를 따라잡았고 그 둘 사이엔 어마어마한 전투가 벌어졌대. "
            "총과 대포가 사방에서 마구 발사되자 위험을 느낀 윌리는 동굴 속으로 피신했는데 영국군이 폭격으로 동굴을 무너뜨렸고 윌리는 그 속에 영원히 갇혀 버렸대.",
            "조용하지만 자존감이 강한 20대 남자 - 술을 사랑하는 사람이 있습니다. 하지만 저는 술을 사랑하는 그 사람도 사랑하고 싶어요. 저희 바텐더는, 형태가 없는 일입니다. "
            "그렇기 때문에 높은 이상을 갖지 않으면, 스스로가 무너지고 말죠. 일이라면 속일 수 있지만, 자신의 인생은 속일 수 없으니까요. 그래도 저는, 바텐더란 일과 만나서 다행이라고 생각 합니다. "
            "바텐더라는 직업이 아니라 바텐더라는, 인생을 선택해서…",
            "감정 콘트롤이 능수능란한 열혈 남자 - 맞아!난코드크라운을되찾기위해베르제브몬과디지털월드에서목숨을걸고싸워왔어하지만다크나이트몬은일곱왕국의데스제너럴들에게다크니스로더를나눠줬지"
            "그덕에녀석들은디지크로스의능력을얻었어게다가도르빅크몬은상대와싸울때지형을마음대로다룰수있는무시무시한녀석이야.",
            "비열한 악역 - 놈의 얼굴을 확인한 순간 저야말로 놈의 살점을 한 점 한 점 발라 죽이고 싶었습니다. 하지만 놈을 그대로 죽여 버리면 양백과 동진을 잡을 수 없습니다. "
            "설령 각시탈을 죽인다 해도 양백과 동진이 살아 있다면 저들의 민족정신이 여전히 살아 숨 쉬게 될 것입니다. 그 싹을 뿌리 채 뽑아, 서구열강의 식민지로 전락해버릴 조선을 기꺼이 구제해준 천황 폐하의 은덕에 감흥 하도록 만들어야합니다",
            "바보와 천재를 오고가는 코믹 캐릭터 - 거실에 보일러 좀 틀어야겠던데 어이쿠! 그렇게 벗고 계셔서야 감기 걸리실까봐 내가 다 걱정이 되더군요. 크크 유치하다고요? "
            "내가 원래 유치한 걸로는 아시아 최고 아닙니까? 유치해서 유치원 다녔고, 유치하다고 유치장 갈 뻔 했습니다. 좋아하는 시인 유치환, 좋아하는 극작가 유치진, 좋아하는 꽃 유채꽃, 그밖에 여러 가지 다양하게 있지요 흐흐",
            "(30대 중반여자 / 살인혐의로 취조를 받는 중산층 여성)\n나는 리처드를 사랑했어요. 둘이서 장미를 가꾸는 것만으로도 행복했죠. 하지만 그걸 안 남편은 미친 듯이 화를 냈죠. 죽일 것 같았어요. "
            "그래서 나도 모르게 그만. 그래도 리처드는 나를 도와줄 거라고 믿었어요. 하지만, 그는 경찰에 알리겠다며, 자수하라고 말했죠. 사실은 날 사랑하지 않았던 거예요.",
            "(10대 중반 여자 / 깜찍한 면이 있는 여고생)\n헤헤... 응. 난 집에서도 큰언니고 쭉 혼자서 잘난 체 하느라 이렇게 응석부리고 싶지 않아. 어머 진짜야. 나 못 믿어? 믿지? 하하하...",
            "(20대 중반 여자 / 카리스마가 있는 여자 마법사)\n사신 카디스는 이 세상에 살아있는 모든 영혼을 증오한다. 망자의 땅으로 변한 로도스에 무슨 번영이 있을 수 있지? "
            "아니면 그게 바로, 네크로맨서가 되고자 하는 당신이 바라는 건가? 후후... 아마 당신 뜻대로 되진 않을걸?",
            "(30-40대 남자)\n이러지마 제발! 난 뭐 너한테 불만이 없는 줄 아니? 처음부터 너의 그 고급취향에 짓 눌려 내가 숨이나 제대로 쉬고 산 줄 알아? "
            "이 여자는 내게 너무 힘들다. 그렇게 생각한 적이 한 두 번이 아니었어. 두둑하지 못한 원고료를 내밀면서 끝없는 자격지심에 시달리게 한 사람이 누군데...",
            "(20대 남자)\n금품 갈취라니요? 양평가서 밥 한번 얻어먹은 것 밖에 없다니까요. (사이) 제 얼굴 한번 보세요. 이 얼굴로 제대로 제비 노릇이나 했겠습니까? "
            "제비도 급이 있는데 내 얼굴로는 하급제비도 못 된다니까요. 그래서 그만 뒀다구요. 정말이예요. 재수 없이 첫 번째 만난 여자가 형사 마누라라서 이 모양이 됐지만.",
            "(1인칭 해설 남자)\n형은 울고 있었다. 그 눈물을 보는 순간, 온 몸에서 맥이 쭉 빠져나가는 것을 느꼈다. 뭐랄까. 형의 울음은 어설펐다. "
            "마치 한번도 울어본 적이 없어서 어떻게 울어야 하는지를 모르는 사람처럼 형은 어설프게 울고 있었다. 슬픔을 안으로 삭이면서 사력을 다해 울음을 참아내고 있는 것처럼 보였다. "
            "(사이) 그러고 보니 형이 우는 모습을 나는 본 적이 없었다.",
            "(해설 남자)\n선거 때까지만 해도 일의 진전은 아주 낙관적이었습니다. 서명자중 거의 대부분이 다시 국회로 진출할 것으로 전망되었습니다. "
            "그리고 그 전망은 적중하였습니다. 그런 상황이면 그해 가을이면 법안은 국회를 통과했어야 마땅합니다. 그러나 96년 부패방지법은 국회의 문턱도 넘지 못하고 맙니다.",
            "(옆에 슬그머니 다가와) 아이고, 드디어 찾았네. 우리 고객님! (입 막으며 협박) 다치기 싫으면 가만있어라. (살벌)니가 감히 나한테 총질까지 해놓고 튀어."
            "그러고도 무사할 줄 알았냐? 그려도 면접 본다는 거 참말이었네? (섬뜩하게) 근데 어쩌냐이~ 난 널 쉽게 놔줄 마음이 없는디. "
            "(웃음) (때리는 소리) 넌 순서가 틀려먹었어. 면접에 가고 싶으면 돈을 가지고 오란 말이여 (E.주먹 날리는) 난 말이여. 고객을 상대할 때 딱 두 가지만 봐."
            "갚을 돈이 있는가?, 갚을 의지가 있는가? 근디 넌 돈도 없고, 뺀질뺀질 도망가느 게 갚을 의지도 없어 보여. 그럴 땐 이렇게 패서 (E.주먹 날리는) 의지를 만들어줘야제."
            "(비웃음) 이게 아주 지대로 미쳐버렸네. 좋아. 그렇게 갚는다고 큰소리를 치니께. 내가 특별히 기회를 주것어."
            " (석구에게 종이 내민다) 여기 이름 쓰고 지장 찍어. 뭐긴 뭐여. 대출서류지. 대출 받아서 연체 이자부터 갚어. 갚을 의지가 있담서. 얼른 이름 써. "
            "(석구가 억지로 서명하자 서류 챙기며) 그려. 오늘까지 연체 된 건 이걸로 됐고. (종이 또 내미는) 여기 하나 더 사인혀. 생명보험이여. 여기 서명만 허면 돼."
            " 불법이 뭔디? 세상은 나 같이 이기는 놈이 합법이고 너처럼 당하는 놈이 결국 불법이 되는 것이여. 더 맞기 전에 얼른 써라이~ "
            "몸이라도 성해야지 돈을 갚을 거 아니여. 아님 결국 이 보험으로 갚을래나. 하하하...(악마처럼 웃으며) 그려. (석구에게) 여기 수익자 양병헌 보이제. "
            "이제 니 몸은 내 담보가 된 거여. 보험료도 대출금으로 정산 될 거니께. 그렇게 알고. (뺨 때리며 철썩) 야! (철썩) 야! 결국 이럴 거 뭐 하러 도망치고 지랄이여. 괜히 힘만 들게."
            " (웃으며) 또 보자이~ (KBS무대 '면접', 면접장까지 쫒아온 사채업자가 석구에게 폭력과 협박을 통해 대출서류에 서명받는데...)",
            "하나님 아버지. 부디 저를 가엾게 여기시고... 시험에 들게 하지 마옵시며... 주뻥의 주둥이를 찢어발겨 주시고... (좋아하는 성당오빠 나타나자) 평안과 안식만을 주시며... "
            "친구의 시름을 함께 짊어지게 하시고... 어머? 오빠? 언제 오셨어요. 온지도 몰랐네. 콜록콜록... 아니에요, 괜찮아요. "
            "몸이 약한 덕분에 건강이 얼마나 감사한 일인지 늘 가슴 깊이 느낄 수 있는 걸요. "
            "(마음속) 내 머리를 쓰다듬는다. 아... 손끝에서 나는 비누향기... 너무... 좋다. (정신들며) 아! 네 오빠. 저 근데... 오늘 저녁에...네? 제가... 뭘.. 우리! 네... 그렇죠 당연히 안 잊었죠."
            " (성당오빠 나가자) 뭐지? 저녁에 뭐지? 내가 나 모르는 사이에 데이트 신청이라도 했나? 아닌가... 지금 오빠가 나한테 데이트 신청 한건가?"
            " 뭐야. 부끄러워서 은근슬쩍 저런 식으로 데이트 신청 한거야? 어우 깍쟁이. 난 몰라. (윤발이 나타나자) 아놔 기분 좋았는데... 에이씨. "
            "아 또 왜 뭐? 지랄, 맡겨 놨냐? 그거 뽑아버리기 전에 조용해라. 누님이 지금 기분이 아주 좋으시다. 망치지 말고 가라. 좋은 기회잖아?"
            " 아... 진짜. 야 주뻥. 그래 내가 진짜 너네 엄마 생각해서 진지하게 고민해 봤는데 그래서 웬만하면 들어주려고 했는데... 내가... 아무래도 오늘 첫 키스, 아니 첫 데이트를 하게 될 거 같아. "
            "다른 남자의 여자가 맘에도 없는 동네친구랑 결혼을 할 수 없잖아? 안 그래? 다 들었으면서 뭘 또 물어."
            " 그럼 이 얘기는 오늘 이 시간부로 그만 하는 걸로... 그리고 울 엄마한테 얘기하면 넌 뒤지시구요. (콧노래부르며 나간다) 기분도 좋은데 동자랑 스파링이나 한판 떠야겠다."
            "(KBS무대 '나의 첫 번째 결혼식', 동네친구 윤발이 자신의 어머니를 위해 가짜 결혼식을 제안하지만 정작 수연은 성당오빠에게 관심이 있고.."]  # 랜덤단문 배열입니다.

        randomNum = random.randrange(0, len(shortscript))  # 0 ~ 랜덤단문 배열 크기 중 랜덤숫자를 지정합니다.
        print("랜덤수 값 :" + str(randomNum))
        print(shortscript[randomNum])
        await client.send_message(message.channel,
                                  embed=discord.Embed(description=shortscript[randomNum]))  # 랜덤 단문을 메시지로 출력합니다.


    if message.content.startswith('~공채대본 남'):
        channel = message.channel
        embed = discord.Embed(
            title='공채기출대본 남자',
            description='```KBS공채 기출 대본 1차```',
            colour=discord.Colour.green()
        )

        embed.set_footer(text='')
        embed.add_field(name='남자 1차 시험 문제', value='(옆에 슬그머니 다가와) 아이고, 드디어 찾았네. 우리 고객님! (입 막으며 협박) 다치기 싫으면 가만있어라. (살벌)니가 감히 나한테 총질까지 해놓고 튀어. '
                                                  '그러고도 무사할 줄 알았냐? 	그려도 면접 본다는 거 참말이었네? 	(섬뜩하게) '
                                                  '근디 어쩌냐이~ 난 널 쉽게 놔줄 마음이 없는디.(웃음) (때리는 소리) 넌 순서가 틀려먹었어. 면접에 가고 싶으면 돈을 가지고 오란 말이여(E.주먹 날리는) '
                                                  '난 말이여. 고객을 상대할 때 딱 두 가지만 봐. 갚을 돈이 있는가?, 갚을 의지가 있는가? 근디 넌 돈도 없고, 뺀질뺀질 도망가는 게 갚을 의지도 없어 보여. '
                                                  '그럴 땐 이렇게 패서 (E.주먹 날리는) 의지를 만들어줘야제. '
                                                  '(비웃음) 이게 아주 지대로 미쳐버렸네. 좋아. 그렇게 갚는다고 큰소리를 치니께. 내가 특별히 기회를 주것어. (석구에게 종이 내민다) 여기 이름 쓰고 지장 찍어. '
                                                  '뭐긴 뭐여. 대출서류지. 대출 받아서 연체 이자부터 갚어. 갚을 의지가 있담서. 얼른 이름 써. (석구가 억지로 서명하자 서류 챙기며) '
                                                  '그려. 오늘까지 연체 된 건 이걸로 됐고. (종이 또 내미는) 여기 하나 더 사인혀. 생명보험이여. '
                                                  '여기 서명만 허면 돼. 불법이 뭔디? 세상은 나 같이 이기는 놈이 합법이고 '
                                                  '너처럼 당하는 놈이 결국 불법이 되는 것이여. 더 맞기 전에 얼른 써라이~ 몸이라도 성해야지 돈을 갚을 거 아니여. 아님 결국 이 보험으로 갚을래나. '
                                                  '하하하...(악마처럼 웃으며) 그려. (석구에게) 여기 수익자 양병헌 보이제. 이제 니 몸은 내 담보가 된 거여. '
                                                  '보험료도 대출금으로 정산 될 거니께. 그렇게 알고. (뺨 때리며 철썩) 야! (철썩) 야! 결국 이럴 거 뭐 하러 도망치고 지랄이여. 괜히 힘만 들게. (웃으며) 또 보자이~ '
                                                  '(KBS무대 ‘면접’, 면접장까지 쫒아온 사채업자가 석구에게 폭력과 협박을 통해 대출서류에 서명받는데...)', inline=False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith('~공채대본 여'):
        channel = message.channel
        embed = discord.Embed(
            title='공채기출대본 여자',
            description='```KBS공채 기출 대본 1차```',
            colour=discord.Colour.green()
        )

        embed.set_footer(text='')
        embed.add_field(name='여자 1차 시험 문제', value='하나님 아버지. 부디 저를 가엾게 여기시고... 시험에 들게 하지 마옵시며... '
                                                  '주뻥의 주둥이를 찢어발겨 주시고...(좋아하는 성당오빠 나타나자)평안과 안식만을 주시며... '
                                                  '친구의 시름을 함께 짊어지게 하시고...어머? 오빠? 언제 오셨어요. 온지도 몰랐네. 콜록콜록...아니에요, 괜찮아요. '
                                                  '몸이 약한 덕분에 건강이 얼마나 감사한 일인지 늘 가슴 깊이 느낄 수 있는걸요. (마음속)내 머리를 쓰다듬는다. '
                                                  '아... 손끝에서 나는 비누향기... 너무... 좋다.(정신들며) 아! 네 오빠. 저 근데... 오늘 저녁에...네? 제가... 뭘.. 우리! 네... 그렇죠 당연히 안 잊었죠. '
                                                  '(성당오빠 나가자)뭐지? 저녁에 뭐지? 내가 나 모르는 사이에 데이트 신청이라도 했나? 아닌가... '
                                                  '지금 오빠가 나한테 데이트 신청 한건가? 뭐야. 부끄러워서 은근슬쩍 저런 식으로 데이트 신청 한거야? 어우 깍쟁이. '
                                                  '난 몰라. (윤발이 나타나자) 아놔 기분 좋았는데... 에이씨. 아 또 왜 뭐? 지랄, 맡겨 놨냐? 그거 뽑아버리기 전에 조용해라. '
                                                  '누님이 지금 기분이 아주 좋으시다. 망치지 말고 가라. 좋은 기회잖아? 아... 진짜. 야 주뻥. '
                                                  '그래 내가 진짜 너네 엄마 생각해서 진지하게 고민해 봤는데 그래서 웬만하면 들어 주려고 했는데...내가... 아무래도 오늘 첫 키스, 아니 첫 데이트를 하게 될 거 같아. '
                                                  '다른 남자의 여자가 맘에도 없는 동네친구랑 결혼을 할 순 없잖아? 안 그래? 다 들었으면서 뭘 또 물어. '
                                                  '그럼 이 얘기는 오늘 이 시간부로 그만 하는 걸로... 그리고 울 엄마한테 얘기하면 넌 뒤지시구요. '
                                                  '(콧노래부르며 나간다) 기분도 좋은데 동자랑 스파링이나 한판 떠야겠다. '
                                                  '(KBS무대 ‘나의 첫 번째 결혼식’, 동네친구 윤발이 자신의 어머니를 위해 가짜 결혼식을 제안하지만 정작 수연은 성당오빠에게 관심이 있고..)', inline=False)

        await client.send_message(channel, embed=embed)



    if message.content.startswith('~주사위'):

        randomNum = random.randrange(1, 7)  # 1~6까지 랜덤수
        print(randomNum)
        if randomNum == 1:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':one:'))
        if randomNum == 2:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':two:'))
        if randomNum == 3:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':three:'))
        if randomNum == 4:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':four:'))
        if randomNum == 5:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':five:'))
        if randomNum == 6:
            await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + ':six: '))

    if message.content.startswith("~날씨"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location + '날씨')
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
        print(url)

    if message.content.startswith("~날씨"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location + '날씨')
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        todayBase = bsObj.find('div', {'class': 'main_info'})

        todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
        todayTemp = todayTemp1.text.strip()  # 온도
        print(todayTemp)

        todayValueBase = todayBase.find('ul', {'class': 'info_list'})
        todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
        todayValue = todayValue2.text.strip()  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        print(todayValue)

        todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
        todayFeelingTemp = todayFeelingTemp1.text.strip()  # 체감온도
        print(todayFeelingTemp)

        todayMiseaMongi1 = bsObj.find('div', {'class': 'sub_info'})
        todayMiseaMongi2 = todayMiseaMongi1.find('div', {'class': 'detail_box'})
        todayMiseaMongi3 = todayMiseaMongi2.find('dd')
        todayMiseaMongi = todayMiseaMongi3.text  # 미세먼지
        print(todayMiseaMongi)

        tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
        tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
        tomorrowTemp2 = tomorrowTemp1.find('dl')
        tomorrowTemp3 = tomorrowTemp2.find('dd')
        tomorrowTemp = tomorrowTemp3.text.strip()  # 오늘 오전,오후온도
        print(tomorrowTemp)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowMoring1 = tomorrowAreaBase.find('div', {'class': 'main_info morning_box'})
        tomorrowMoring2 = tomorrowMoring1.find('span', {'class': 'todaytemp'})
        tomorrowMoring = tomorrowMoring2.text.strip()  # 내일 오전 온도
        print(tomorrowMoring)

        tomorrowValue1 = tomorrowMoring1.find('div', {'class': 'info_data'})
        tomorrowValue = tomorrowValue1.text.strip()  # 내일 오전 날씨상태, 미세먼지 상태
        print(tomorrowValue)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
        tomorrowAfter1 = tomorrowAllFind[1]
        tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
        tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
        tomorrowAfterTemp = tomorrowAfter3.text.strip()  # 내일 오후 온도
        print(tomorrowAfterTemp)

        tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
        tomorrowAfterValue = tomorrowAfterValue1.text.strip()

        print(tomorrowAfterValue)  # 내일 오후 날씨상태,미세먼지

        embed = discord.Embed(
            title=learn[1] + ' 날씨 정보',
            description=learn[1] + '날씨 정보입니다.',
            colour=discord.Colour.gold()
        )
        embed.add_field(name='현재온도', value=todayTemp + '˚', inline=False)  # 현재온도
        embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
        embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        embed.add_field(name='현재 미세먼지 상태', value=todayMiseaMongi, inline=False)  # 오늘 미세먼지
        embed.add_field(name='오늘 오전/오후 날씨', value=tomorrowTemp, inline=False)  # 오늘날씨 # color=discord.Color.blue()
        embed.add_field(name='**----------------------------------**', value='**----------------------------------**',
                        inline=False)  # 구분선
        embed.add_field(name='내일 오전온도', value=tomorrowMoring + '˚', inline=False)  # 내일오전날씨
        embed.add_field(name='내일 오전날씨상태, 미세먼지 상태', value=tomorrowValue, inline=False)  # 내일오전 날씨상태
        embed.add_field(name='내일 오후온도', value=tomorrowAfterTemp + '˚', inline=False)  # 내일오후날씨
        embed.add_field(name='내일 오후날씨상태, 미세먼지 상태', value=tomorrowAfterValue, inline=False)  # 내일오후 날씨상태

        await client.send_message(message.channel, embed=embed)

    if message.content.startswith('~골라'):
        choice = message.content.split(" ")
        choicenumber = random.randint(1, len(choice) - 1)
        choiceresult = choice[choicenumber]
        await client.send_message(message.channel, choiceresult)

    if message.content.startswith("~사다리타기"):
        await client.send_message(message.channel, embed=discord.Embed(description='사다리타기를 시작합니다!'))
        team = message.content[7:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel,
                                      embed=discord.Embed(description=person[i] + "---->" + teamname[i]))

    if message.content.startswith('~제비뽑기'):
        channel = message.channel
        embed = discord.Embed(
            title='제비뽑기',
            description='각 번호별로 번호를 지정합니다.',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='끝')

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]
        print(Text.strip())  # 입력한 명령어

        number = int(Text)

        List = []
        num = random.randrange(0, number)
        for i in range(number):
            while num in List:  # 중복일때만
                num = random.randrange(0, number + 1)  # 다시 랜덤수 생성

            List.append(num)  # 중복 아닐때만 리스트에 추가
            embed.add_field(name=str(i + 1) + '번째', value=str(num + 1), inline=True)

        print(List)
        await client.send_message(channel, embed=embed)

    if message.content.startswith('~이미지'):

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]
        print(Text.strip())  # 입력한 명령어

        randomNum = random.randrange(0, 40)  # 랜덤 이미지 숫자

        location = Text
        enc_location = urllib.parse.quote(location)  # 한글을 url에 사용하게끔 형식을 바꿔줍니다. 그냥 한글로 쓰면 실행이 안됩니다.
        hdr = {'User-Agent': 'Mozilla/5.0'}
        # 크롤링 하는데 있어서 가끔씩 안되는 사이트가 있습니다.
        # 그 이유는 사이트가 접속하는 상대를 봇으로 인식하였기 때문인데
        # 이 코드는 자신이 봇이 아닌것을 증명하여 사이트에 접속이 가능해집니다!
        url = 'https://search.naver.com/search.naver?where=image&sm=tab_jum&query=' + enc_location  # 이미지 검색링크+검색할 키워드
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")  # 전체 html 코드를 가져옵니다.
        # print(bsObj)
        imgfind1 = bsObj.find('div', {'class': 'photo_grid _box'})  # bsjObj에서 div class : photo_grid_box 의 코드를 가져옵니다.
        # print(imgfind1)
        imgfind2 = imgfind1.findAll('a', {'class': 'thumb _thumb'})  # imgfind1 에서 모든 a태그 코드를 가져옵니다.
        imgfind3 = imgfind2[randomNum]  # 0이면 1번째사진 1이면 2번째사진 형식으로 하나의 사진 코드만 가져옵니다.
        imgfind4 = imgfind3.find('img')  # imgfind3 에서 img코드만 가져옵니다.
        imgsrc = imgfind4.get('data-source')  # imgfind4 에서 data-source(사진링크) 의 값만 가져옵니다.
        print(imgsrc)
        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed.set_image(url=imgsrc)  # 이미지의 링크를 지정해 이미지를 설정합니다.
        embed.add_field(name='검색 : ' + Text, value='링크 : ' + imgsrc, inline=False)
        await client.send_message(message.channel, embed=embed)  # 메시지를 보냅니다.



    if message.content.startswith('~커맨드') and not message.content.startswith("~커맨드삭제") and not message.content.startswith("~커맨드목록"):
        learn = message.content[5:]
        learn1 = learn.split("/")
        file = openpyxl.load_workbook('기억.xlsx')
        sheet = file.active
        i = 1
        while True:
            if sheet["C" + str(i)].value == str(message.server.id) and sheet["A" + str(i)].value == learn1[0] and sheet["B" + str(i)].value == learn1[1]:
                await client.send_message(message.channel, "당신은 이미 동일한 커맨드를 추가시켰습니다.")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = learn1[0]
                sheet["B" + str(i)].value = learn1[1]
                sheet["C" + str(i)].value = str(message.server.id)
                sheet["D" + str(i)].value = str(message.author.id)
                file.save("기억.xlsx")
                await client.send_message(message.channel, "커맨드가 추가되었습니다.")
                break
            i += 1

    if message.content.startswith(""):
        msg = message.content[0:]
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        i = 1
        list = []
        while True:
            if sheet["A" + str(i)].value == msg and sheet["C" + str(i)].value == str(message.server.id):
                list.append(sheet["B" + str(i)].value)
            if sheet["A" + str(i)].value == None:
                if len(list) == 0:
                    break
                else:
                    num = random.randint(0, len(list)-1)
                    await client.send_message(message.channel, str(list[num]))
                    break
            i += 1

    if message.content.startswith("~커맨드삭제"):
        msg = message.content[7:]
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == msg and sheet["C" + str(i)].value == str(message.server.id) and sheet["D" + str(i)].value == str(message.author.id):
                sheet["A" + str(i)].value = "/"
                sheet["B" + str(i)].value = "/"
                sheet["C" + str(i)].value = "/"
                sheet["D" + str(i)].value = "/"
                file.save('기억.xlsx')
            if sheet["A" + str(i)].value == None:
                await client.send_message(message.channel, "해당 커맨드가 삭제되었습니다.")
                break
            i += 1

    if message.content.startswith("~커맨드목록"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        i = 1
        list = []
        while True:
            if sheet["C" + str(i)].value == str(message.server.id) and sheet["D" + str(i)].value == str(message.author.id):
                list.append(sheet["A" + str(i)].value + " / " + sheet["B" + str(i)].value)
            if sheet["A" + str(i)].value == None:
                if len(list) == 0:
                    await client.send_message(message.channel, "학습한 단어가 없습니다.")
                    break
                else:
                    txt = ""
                    for a in range(0, len(list)):
                        txt = txt + "\n" + list[a]
                    embed = discord.Embed(description=txt)
                    await client.send_message(message.channel, embed=embed)
                    break
            i += 1

    if message.content.startswith("~!전체공지"):
        msg = message.content[6:]
        for server in client.servers:
            for channel in server.channels:
                try:
                    await client.send_message(channel, msg)
                    break
                except:
                    a = 1



client.run('NTIzNzk4OTY1NTE5NjQ2NzM0.DvexaA.gAo6iSiJdc57UwuWwQY7DC_0qSA')